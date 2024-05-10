import requests
import sys
import os
import json
import openpyxl
import dateparser
from datetime import timedelta
from urllib.parse import urlencode


os.chdir(os.path.dirname(os.path.realpath(sys.argv[0])))
re = json.load(open("conf.json"))
USER = re["user"]
STORY = re["story"]
PROJECT = re["project"]
COOKIE = re["cookie"]
ON = re["on"]
OFF = re["off"]


def calDate(period, remain):
    dd, dt = int(period / 8), period % 8
    remain -= dt
    if remain <= 0:
        dd += 1
        remain += 8
    return dd, remain


def calWork(now, day):
    for i in range(day):
        flag = True
        while flag:
            now += timedelta(days=1)
            sts = now.strftime("%Y%m%d")
            if now.weekday() >= 5:
                flag = sts not in ON  # 周末上班，不加
            else:
                flag = sts in OFF  # 工作日休息，加一天
    return now


class tapdTask():
    headers = {"user-agent":  "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"}
    datas = []
    allTaskUrl = "https://www.tapd.cn/api/entity/tasks/task_list_by_condition"
    workId = ""
    oldData = []

    def __init__(self, story, cookie) -> None:
        self.story = story
        self.headers["cookie"] = cookie
        for i in cookie.split(";"):
            if "cloud_current_workspaceId" in i:
                self.workId = i.split("=")[1]
        self.taskUrl = "https://www.tapd.cn/%s/prong/tasks/quick_add_task/%s?is_from_story_view=true" % (self.workId, story)
        self.doneUrl = "https://www.tapd.cn/%s/prong/tasks/changeStatus?new_status=done&objid=" % self.workId
        self.load()

    def load(self):
        try:
            with open(self.story+".json", 'r') as f:
                self.oldData = json.load(f)
        except FileNotFoundError:
            pass

    def save(self):
        with open(self.story+".json", 'w') as f:
            json.dump(self.oldData, f, indent=4)

    def read(self, fp, sheet, project, user, startDay):
        book = openpyxl.load_workbook(fp)
        sheet = book[sheet]
        for i in sheet.merged_cells.ranges.copy():
            cell_start = i.start_cell
            sheet.unmerge_cells(range_string=i.coord)
            for row_index, col_index in i.cells:
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = cell_start.value
        datas = [[cell.value for cell in row] for row in sheet]
        spendi, useri, starti = 0, 0, 0
        for i, title in enumerate(datas[0]):
            if "花费" in title:
                spendi = i
            elif "负责人" in title:
                useri = i
            elif "开始日期" == title:
                starti = i
        if spendi == 0:
            print("缺少花费时间列")
            exit()

        res = []
        currentDay = dateparser.parse(startDay)
        currentStr = currentDay.strftime("%Y-%m-%d")
        remain = 8
        days = {}
        if starti:
            for data in datas[1:]:
                if data[starti] and data[spendi]:
                    sts = dateparser.parse(data[starti]).strftime("%Y-%m-%d")
                    if sts in days:
                        days[sts] += data[spendi]
                    else:
                        days[sts] = data[spendi]

        for data in datas[1:]:
            re = [project, data[spendi], user, currentStr, ""]
            if re[1]:
                for i in range(spendi):
                    v = data[i]
                    if v:
                        re[0] += ":"+v
                if useri and data[useri]:
                    re[2] = data[useri]

                if starti and data[starti]:
                    st = dateparser.parse(data[starti])
                    re[3] = st.strftime("%Y-%m-%d")
                    st = calWork(st, int(re[1] / 8))
                    if re[1] % 8:
                        st = calWork(st, 1)
                    re[4] = st.strftime("%Y-%m-%d")
                    res.append(re)
                    continue

                if currentStr in days:
                    dd = days.pop(currentStr)
                    dd, remain = calDate(dd, remain)
                    # print("date", dd, remain)
                    currentDay = calWork(currentDay, dd)
                    re[3] = currentDay.strftime("%Y-%m-%d")

                dd, remain = calDate(re[1], remain)
                # print(dd, remain)
                if remain == 8 and dd > 0:
                    currentDay = calWork(currentDay, dd-1)
                    re[4] = currentDay.strftime("%Y-%m-%d")
                    currentDay = calWork(currentDay, 1)
                else:
                    currentDay = calWork(currentDay, dd)
                    re[4] = currentDay.strftime("%Y-%m-%d")
                currentStr = currentDay.strftime("%Y-%m-%d")
                res.append(re)
        self.datas = res
        return self

    def create(self):
        for i in self.datas:
            self.createOne(i)

    def createOne(self, d):
        print("开始添加: " + d[0])
        data = {
            "data[Task][name]": d[0],
            "data[Task][effort]": d[1],
            "data[Task][owner]": d[2]+";",
            "data[Task][begin]": d[3],
            "data[Task][due]": d[4]
        }
        response = requests.post(self.taskUrl, data=urlencode(data), headers={**self.headers,  "content-type": "application/x-www-form-urlencoded; charset=UTF-8"})
        if response.status_code != 200:
            print("添加失败:", response.text)
            return response.text
        return "完成"

    def taskIds(self):
        query = {"workspace_ids": [self.workId], "page_count": 1, "story_id": self.story, "from": "story_relate_task"}
        response = requests.post(self.allTaskUrl, json=query, headers=self.headers)
        if response.status_code != 200:
            print("查询失败:", response.text)
            return {}
        re = {i["title"]: i["id"] for i in response.json()["data"]["list_excludes"]["checked_list_with_name"]}
        if len(self.oldData) > 0:
            old = self.oldData.copy()
            for k, v in re.items():
                if v in old:
                    re[k] = "完成"
                    old.remove(v)
                    if len(old) == 0:
                        break
        return re

    def done(self, id):
        if id and id != "完成":
            if requests.get(self.doneUrl + id, headers=self.headers).status_code == 200:
                self.oldData.append(id)
                return "完成"
        return "失败"
